import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def workbook_process(fileName):
    wb = xl.load_workbook(fileName)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    vaules = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(vaules)
    sheet.add_chart(chart, 'E2')

    wb.save(fileName)

workbook_process('transactions.xlsx')


